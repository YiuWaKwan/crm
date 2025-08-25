import re
import time

from django.http import HttpResponse, StreamingHttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt

from lib.ModuleConfig import ConfAnalysis
import os
import mySqlUtil
import json
from lib.DateEncoder import DateEncoder

from lib.FinalLogger import getLogger

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# 初始化logger
loggerFIle = 'log/friendViews.log'
logger = getLogger(loggerFIle)
# 初始化config
configFile = 'conf/moduleConfig.conf'
confAllItems = ConfAnalysis(logger, configFile)

downloadFilePath = confAllItems.getOneOptions('file','downloadFilePath')
def chatHistory(request):
    oper_id = request.session['oper_id']
    oper_main_wx = request.session['oper_main_wx']
    joinStr = "','"
    mainWxStr = "'" + joinStr.join(oper_main_wx) + "'"
    sql="select CONCAT(wx_name,'(',wx_login_id,')') NAME,wx_id VALUE from wx_account_info where 1=1 "
    sql = sql + """ and wx_id in(%s)""" % (mainWxStr)

    wxMainList = mySqlUtil.getData(sql)
    wxMainListRet = [i for i in wxMainList]

    return render(request, "chatHistory.html",
                  {'wxMainList': list(wxMainListRet)})

@csrf_exempt
def chatHistoryData(request):
    ret={}
    try:
        if request.method == 'GET':
            type=request.GET.get('type')
        else:
            type=request.POST.get('type')
        if(type == 'getFriendList'):
            ret=getFriendList(request)
    except(Exception) as e:
         logger.error(e)
    return HttpResponse(json.dumps(ret, ensure_ascii=False, cls=DateEncoder))

def getFriendList(request):
    ret={}
    try:
        wxMainIdSearch = request.GET.get('wxMainIdSearch')
        wxIdSearch = request.GET.get('wxIdSearch')
        friendNameSearch = request.GET.get('friendNameSearch')
        isGroup = request.GET.get('isGroup')
        pageSize = request.GET.get("pageSize")
        pageIndex = request.GET.get("pageIndex")

        oper_id=request.session['oper_id']
        sql = """select b.wx_main_id,b.wx_id,b.remark,b.last_chart_time,b.last_chart_content 
                    from wx_friend_rela b where 1=1 and b.last_chart_time is not null  """
        if wxMainIdSearch  and wxMainIdSearch.strip():
            sql=sql+" and wx_main_id = '"+wxMainIdSearch+"' "
        if wxMainIdSearch  and wxIdSearch.strip():
            sql = sql + " and wx_id like '%"+wxIdSearch+"%'"
        if wxMainIdSearch  and friendNameSearch.strip():
            sql = sql + " and remark like '%"+friendNameSearch+"%'"
        if isGroup and isGroup=='1':
            sql = sql + " and b.wx_id like '%@chatroom'"
        if isGroup and isGroup=='0':
            sql = sql + " and b.wx_id not like '%@chatroom'"
         # logger.info(sql)
        ret = mySqlUtil.getDataByPage(sql,pageIndex,pageSize)
    except(Exception) as e:
        logger.error(e)
    return ret
def exportChatHistoryData(request):
    ret=None
    try:
        wxMainId = request.GET.get('wxMainId')
        wxIds = request.GET.get('wxIds')
        startDate = request.GET.get('startDate')
        endDate = request.GET.get('endDate')
        #wxMainId='wxid_sr2end98s8th22'
        #wxIds='4509474861@chatroom,shijianxiong'
        dateRang=''
        if startDate:
            dateRang+=" and send_time >= date_format('"+startDate+"', '%Y-%c-%d %H:%i:%s')"
        if endDate:
            dateRang+=" and send_time <= date_format('"+endDate+"', '%Y-%c-%d %H:%i:%s')"
            #获取主号信息
        sql="select wx_login_id,wx_name from wx_account_info where wx_id='%s'" %wxMainId
        wxMainInfo= mySqlUtil.getData(sql)[0]
        wxMainLoginId=wxMainInfo[0]
        wxMainName=wxMainInfo[1]

        # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
        wb = Workbook()
        wsGroup = wb.active
        wsGroup.title ="群"
        wsGroup.append(["主号", "微信群名称", "ID", "时间","联系人","微信号","状态","类型","消息","系统工号"])
        wsFriend = wb.create_sheet(title="好友")
        wsFriend.append(["主号","时间","联系人","微信号","状态","类型","消息","系统工号"])
        hasGroupMsg=False
        hasFriendMsg=False
        wxIds=wxIds.replace(',',"','")
        sql=u"SELECT (select CONCAT(wx_name,'(',wx_login_id,')') from wx_account_info where wx_id='" + wxMainId + "')," \
        u"(select remark from wx_friend_rela where wx_main_id='" + wxMainId + "' and wx_id=c.wx_id)," \
        u"c.wx_id,date_format(c.send_time,'%Y-%c-%d %H:%i:%s'),group_member_name,group_member_id,c.content," \
        u"(case c.send_type when '1' then '发送' else '接收' end)," \
        u"(case c.type when '1' then '文字' when '2' then '图片'  when '3' then '文件'  when '4' then '系统消息' when '5' then '系统消息'  when '6' then '语音'  when '7' then '视频'  else c.type end)," \
        u"(select name from wx_system_operator where oper_id=c.oper_id) " \
        u" FROM wx_chat_info_his c where  c.wx_main_id='" + wxMainId + "' and c.wx_id in('" + wxIds + "')"+dateRang+" order by wx_id asc,send_time asc,createTime asc "
        logger.info(sql)
        chatListdata = mySqlUtil.getData(sql)
        for info in chatListdata:
            wxMain=info[0]
            groupOrWxName=info[1]
            groupOrWxId=info[2]
            sendTime=info[3]
            groupMemberName=info[4]
            groupMemberId=info[5]
            content=info[6]
            sendType=info[7]
            msgType=info[8]
            operName=info[9]
            if '@chatroom' in groupOrWxId:
                wsGroup.append([wxMain,groupOrWxName,groupOrWxId,sendTime,groupMemberName,groupMemberId,sendType,msgType,content,operName])
            else:
                wsFriend.append([wxMain,sendTime,groupOrWxName,groupOrWxId,sendType,msgType,content,operName])

        # 保存
        fileName=wxMainLoginId+"_"+str(round(time.time() * 1000))
        file_path_name=downloadFilePath+fileName+".xlsx"
        wb.save(filename=file_path_name)
        ret=file_download(request,file_path_name,fileName)
    except(Exception) as e:
        logger.exception(e)
        #ret['flagList'] = flagList
    return ret

def exportChatHistoryDataOld(request):
    ret=None
    try:
        wxMainId = request.GET.get('wxMainId')
        wxIds = request.GET.get('wxIds')
        startDate = request.GET.get('startDate')
        endDate = request.GET.get('endDate')
        #wxMainId='wxid_sr2end98s8th22'
        #wxIds='4509474861@chatroom,shijianxiong'
        dateRang=''
        if startDate:
            dateRang+=" and send_time >= date_format('"+startDate+"', '%Y-%c-%d %H:%i:%s')"
        if endDate:
            dateRang+=" and send_time <= date_format('"+endDate+"', '%Y-%c-%d %H:%i:%s')"
            #获取主号信息
        sql="select wx_login_id,wx_name from wx_account_info where wx_id='%s'" %wxMainId
        wxMainInfo= mySqlUtil.getData(sql)[0]
        wxMainLoginId=wxMainInfo[0]
        wxMainName=wxMainInfo[1]

        # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
        wb = Workbook()
        wxIdArr =wxIds.split(',')
        for index in range(len(wxIdArr)):
            wxId=wxIdArr[index]
            #获取好友或群信息
            sql="select remark from wx_friend_rela where wx_main_id='%s' and wx_id='%s'" %(wxMainId,wxId)
            remarkName= mySqlUtil.getData(sql)[0][0]

            sql = u"SELECT date_format(send_time,'%Y-%c-%d %H:%i:%s'),type,content,c.group_member_id ,c.group_member_name ,c.head_picture,c.msgId,c.send_type " \
                  u" FROM wx_chat_info_his c " \
                  u"where c.wx_main_id='" + wxMainId + "' and c.wx_id='" + wxId + "'"+dateRang+" order by send_time asc,createTime asc"
            logger.info(sql)
            chatListdata = mySqlUtil.getData(sql)
            #sheet文件名特殊字符替换
            rstr = r"[\/\\\:\*\?\"\<\>\|]"  # '/ \ : * ? " < > |'
            remarkName = re.sub(rstr, "_", remarkName)  # 替换为下划线
            # 获取当前活跃的worksheet,默认就是第一个worksheet
            if index==0:
                ws = wb.active
                ws.title =remarkName
            else:
                ws = wb.create_sheet(title=remarkName)
            # 设置单元格的值，A1等于6(测试可知openpyxl的行和列编号从1开始计算)，B1等于7
            #ws.cell(row=1, column=1).value = 6
            #ws.cell("B1").value = 7
            ws.append(["时间", "发送人", "接收人", "内容"])
            # 从第2行开始，写入9行10列数据，值为对应的列序号A、B、C、D...
            #for row in range(2, 11):
            #    for col in range(1, 11):
             #       ws.cell(row=row, column=col).value = get_column_letter(col)
            # 可以使用append插入一行数据
            for info in chatListdata:
                sendUser=''
                acceptUser=''
                sendType=info[7]
                if sendType=='1':#发送的消息
                    sendUser=wxMainName
                    if not '@chatroom' in wxId:
                        acceptUser = remarkName
                else:
                    if '@chatroom' in wxId:
                        sendUser=info[4]
                    else:
                        sendUser = remarkName
                        acceptUser= wxMainName
                ws.append([info[0], sendUser,acceptUser,info[2]])
        # 保存
        fileName=wxMainLoginId+"_"+str(round(time.time() * 1000))
        file_path_name=downloadFilePath+fileName+".xlsx"
        wb.save(filename=file_path_name)
        ret=file_download(request,file_path_name,fileName)
    except(Exception) as e:
        logger.exception(e)
        #ret['flagList'] = flagList
    return ret

def file_download(request,file_path_name,file_name):
    def file_iterator(file_name, chunk_size=512):
        with open(file_name,'rb+') as f:
            while True:
                c = f.read(chunk_size)
                if c:
                    yield c
                else:
                    break
    response = StreamingHttpResponse(file_iterator(file_path_name))
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename="{0}"'.format(file_name+".xlsx")
    return response

if __name__ == '__main__':
    exportChatHistoryData(None)